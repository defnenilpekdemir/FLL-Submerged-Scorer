import tkinter as tk
from tkinter import ttk, messagebox
import datetime
import os

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from PIL import Image, ImageTk

# -----------------------------------------------------------------------------
# Görev ve Alt Maddeler Tanımı (Türkçeleştirilmiş)
# -----------------------------------------------------------------------------
MISSIONS_DATA = [
    {
        "code": "Kurulum",
        "title": "Robot ve Ekipman",
        "items": [
            {
                "type": "checkbox",
                "desc": "Robot ve tüm takım ekipmanları tek bir fırlatma alanına ve\n12 inç (305 mm) altına sığar.",
                "points": 20
            }
        ]
    },
    {
        "code": "M01",
        "title": "Mercan Ağacı",
        "items": [
            {
                "type": "checkbox",
                "desc": "Mercan ağacı, mercan ağacı desteğinde asılı",
                "points": 20
            },
            {
                "type": "checkbox",
                "desc": "Mercan ağacının alt kısmı yuvasında",
                "points": 10
            },
            {
                "type": "checkbox",
                "desc": "Mercan tomurcukları yukarı çevrilmiş",
                "points": 20
            }
        ]
    },
    {
        "code": "M02",
        "title": "Köpekbalığı ve Mağara",
        "items": [
            {
                "type": "checkbox",
                "desc": "Köpekbalığı artık mağaraya dokunmuyor",
                "points": 20
            },
            {
                "type": "checkbox",
                "desc": "Köpekbalığı, köpekbalığı yaşam alanında en azından kısmen mat'a dokunuyor",
                "points": 10
            }
        ]
    },
    {
        "code": "M03",
        "title": "Mercan Resifi",
        "items": [
            {
                "type": "checkbox",
                "desc": "Mercan resifi yukarı kaldırıldı, mat'a dokunmuyor",
                "points": 20
            },
            {
                "type": "numeric",
                "desc": "Ev alanı dışında ve mat'a dokunan, dik duran resif segment(ler)inin sayısı",
                "points_per_item": 5
            }
        ]
    },
    {
        "code": "M04",
        "title": "Dalgıç",
        "items": [
            {
                "type": "checkbox",
                "desc": "Dalgıç artık mercan fidanlığına dokunmuyor",
                "points": 20
            },
            {
                "type": "checkbox",
                "desc": "Dalgıç, mercan resifi desteğinde asılı duruyor",
                "points": 20
            }
        ]
    },
    {
        "code": "M05",
        "title": "Fener Balığı",
        "items": [
            {
                "type": "checkbox",
                "desc": "Fener balığı, gemi enkazı içinde kilitli",
                "points": 30
            }
        ]
    },
    {
        "code": "M06",
        "title": "Gemi Enkazı Direği",
        "items": [
            {
                "type": "checkbox",
                "desc": "Gemi enkazının direği tamamen kaldırıldı",
                "points": 30
            }
        ]
    },
    {
        "code": "M07",
        "title": "Hazine Sandığı",
        "items": [
            {
                "type": "checkbox",
                "desc": "Hazine sandığı tamamen kraken yuvasının dışında",
                "points": 20
            }
        ]
    },
    {
        "code": "M08",
        "title": "Yapay Habitat",
        "items": [
            {
                "type": "numeric",
                "desc": "Tamamen düz ve dik duran yapay habitat segment(ler)inin sayısı",
                "points_per_item": 10
            }
        ]
    },
    {
        "code": "M09",
        "title": "Sıra Dışı Karşılaşma",
        "items": [
            {
                "type": "checkbox",
                "desc": "Bilinmeyen yaratık serbest bırakıldı",
                "points": 20
            },
            {
                "type": "checkbox",
                "desc": "Bilinmeyen yaratık, soğuk çıkış alanında (cold seep) en azından kısmen bulunuyor",
                "points": 10
            }
        ]
    },
    {
        "code": "M10",
        "title": "Sarı Bayrak ve Denizaltı",
        "items": [
            {
                "type": "checkbox",
                "desc": "Takımınızın sarı bayrağı indirilmiş",
                "points": 30
            },
            {
                "type": "checkbox",
                "desc": "Denizaltı, açıkça rakip alana daha yakın",
                "points": 10
            }
        ]
    },
    {
        "code": "M11",
        "title": "Ortaya Çıkan Balinalar",
        "items": [
            {
                "type": "numeric",
                "desc": "Ortaya çıkarılan balinaların sayısı",
                "points_per_item": 15
            }
        ]
    },
    {
        "code": "M12",
        "title": "Kril ve Balina",
        "items": [
            {
                "type": "numeric",
                "desc": "Balinanın ağzında en azından kısmen bulunan kril sayısı",
                "points_per_item": 10
            }
        ]
    },
    {
        "code": "M13",
        "title": "Nakliye Rotası",
        "items": [
            {
                "type": "checkbox",
                "desc": "Gemi, yeni nakliye rotasında ve mat'a dokunuyor",
                "points": 20
            }
        ]
    },
    {
        "code": "M14",
        "title": "Numuneler ve Üç Dişli",
        "items": [
            {
                "type": "checkbox",
                "desc": "Su numunesi, su numunesi alanının tamamen dışında",
                "points": 5
            },
            {
                "type": "checkbox",
                "desc": "Deniz tabanı numunesi artık deniz tabanına dokunmuyor",
                "points": 10
            },
            {
                "type": "checkbox",
                "desc": "Plankton numunesi artık yosun ormanına dokunmuyor",
                "points": 10
            },
            {
                "type": "numeric",
                "desc": "Gemi enkazına artık dokunmayan üç dişli (trident) parça(lar)ının sayısı",
                "points_per_item": 15
            }
        ]
    },
    {
        "code": "M15",
        "title": "Araştırma Gemisi",
        "items": [
            {
                "type": "numeric",
                "desc": "Araştırma gemisinin yük bölümünde en azından kısmen bulunan\nnumune(ler)/üç dişli parçası/hazine sandığı sayısı",
                "points_per_item": 5
            },
            {
                "type": "checkbox",
                "desc": "Liman mandalı, araştırma gemisinin halkasına en azından kısmen takılı",
                "points": 20
            }
        ]
    },
    {
        "code": "Hassasiyet Pulu",
        "title": "Precision Tokens",
        "items": [
            {
                "type": "numeric",
                "desc": "Kalan hassasiyet pulu sayısı",
                "points_per_item": 10
            }
        ]
    }
]

class SubmergedScorerApp(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("FLL Submerged Puanlama - KRUSTY KRABS")
        self.geometry("900x700")  # Pencere başlangıç boyutu

        # ------------------ ÜST BÖLÜM (Sabit) ------------------
        top_frame = ttk.Frame(self)
        top_frame.pack(fill=tk.X, side=tk.TOP)

        # 1) Skor etiketi (Solda)
        self.big_score_label = tk.Label(
            top_frame,
            text="Skor: 0",
            fg="white",
            bg="blue",
            font=("Arial", 20, "bold"),
            width=20,
            height=2,
            anchor="w"
        )
        self.big_score_label.pack(side="left", padx=10, pady=10)

        # 2) LOGO ORTADA
        logo_frame = ttk.Frame(top_frame)
        logo_frame.pack(side="left", expand=True, fill="both")

        # Pillow ile resmi yükleyip yeniden boyutlandırma
        try:
            original_image = Image.open("/Users/oguzhankose/PycharmProjects/FLL/logo.jpeg")  # veya tam yol: Image.open("C:/path/to/logo.png")
            resized_image = original_image.resize((100, 100), Image.LANCZOS)  # Örnek: 100x100 piksel
            self.logo_image = ImageTk.PhotoImage(resized_image)
        except Exception as e:
            messagebox.showerror("Logo Hatası", f"logo.png yüklenemedi.\n{e}")
            self.logo_image = None

        if self.logo_image:
            self.logo_label = tk.Label(logo_frame, image=self.logo_image)
            self.logo_label.pack(anchor="center")

        # 3) Sağdaki butonlar (Kaydet ve Temizle)
        button_frame = ttk.Frame(top_frame)
        button_frame.pack(side="right", padx=(0, 20))

        # Temizle butonu
        self.clear_button = ttk.Button(button_frame, text="Skorları Temizle", command=self.clear_scores)
        self.clear_button.pack(side="right", padx=(0, 30), pady=10)

        # Kaydet butonu
        self.save_button = ttk.Button(button_frame, text="Kaydet", command=self.save_score)
        self.save_button.pack(side="right", padx=10, pady=10)

        # ------------------ ALT BÖLÜM (Kaydırılabilir İçerik) ------------------
        bottom_frame = ttk.Frame(self)
        bottom_frame.pack(fill=tk.BOTH, expand=True)

        self.canvas = tk.Canvas(bottom_frame)
        self.canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        self.scrollbar = ttk.Scrollbar(bottom_frame, orient="vertical", command=self.canvas.yview)
        self.scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

        self.canvas.configure(yscrollcommand=self.scrollbar.set)

        # İçerik çerçevesi (kaydırılacak kısım)
        self.scrollable_frame = ttk.Frame(self.canvas)
        self.canvas.create_window((0, 0), window=self.scrollable_frame, anchor="nw")

        # Boyut değişince scrollbar'ı güncelle
        self.scrollable_frame.bind(
            "<Configure>",
            lambda e: self.canvas.configure(scrollregion=self.canvas.bbox("all"))
        )

        # Fare kaydırma
        # Windows/Linux
        self.canvas.bind_all("<MouseWheel>", self._on_mousewheel)
        # Mac OS
        self.canvas.bind_all("<Button-4>", self._on_mousewheel_mac)
        self.canvas.bind_all("<Button-5>", self._on_mousewheel_mac)

        # Değerleri tutacak listeler
        self.checkbox_vars = []
        self.numeric_vars = []

        # Arayüzdeki görevleri ve alanları oluştur
        self.create_widgets()

    def _on_mousewheel(self, event):
        """Windows/Linux tekerleği."""
        self.canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")

    def _on_mousewheel_mac(self, event):
        """MacOS <Button-4> / <Button-5>."""
        if event.num == 4:  # Yukarı kaydır
            self.canvas.yview_scroll(-1, "units")
        elif event.num == 5:  # Aşağı kaydır
            self.canvas.yview_scroll(1, "units")

    def create_widgets(self):
        """Görev listelerini ekrana yerleştirir."""
        title_label = ttk.Label(
            self.scrollable_frame,
            text="FLL Submerged\n Skor Hesaplama",
            font=("Arial", 12, "bold")
        )
        title_label.pack(pady=5)

        # Görevleri ekrana ekle
        for mission_data in MISSIONS_DATA:
            mf = ttk.LabelFrame(
                self.scrollable_frame,
                text=f"{mission_data['code']} - {mission_data['title']}"
            )
            mf.pack(fill=tk.X, padx=5, pady=5)

            for item_data in mission_data["items"]:
                if item_data["type"] == "checkbox":
                    var = tk.BooleanVar(value=False)
                    self.checkbox_vars.append((var, item_data))
                    cb = ttk.Checkbutton(
                        mf,
                        text=item_data["desc"],
                        variable=var,
                        command=self.update_score
                    )
                    cb.pack(anchor=tk.W, padx=10, pady=2)
                elif item_data["type"] == "numeric":
                    f2 = ttk.Frame(mf)
                    f2.pack(anchor=tk.W, padx=10, pady=2, fill=tk.X)

                    lbl = ttk.Label(f2, text=item_data["desc"])
                    lbl.pack(side=tk.LEFT)

                    var = tk.IntVar(value=0)
                    self.numeric_vars.append((var, item_data))

                    entry = ttk.Entry(f2, width=5, textvariable=var)
                    entry.pack(side=tk.LEFT, padx=5)

                    entry.bind("<KeyRelease>", lambda e: self.update_score())
                    entry.bind("<FocusOut>", lambda e: self.update_score())

        # Altta “Toplam Puan”
        sf = ttk.Frame(self.scrollable_frame)
        sf.pack(fill=tk.X, pady=10)

        lbl_score_title = ttk.Label(sf, text="Toplam Puan:", font=("Arial", 12, "bold"))
        lbl_score_title.pack(side=tk.LEFT)

        self.lbl_score_value = ttk.Label(sf, text="0", font=("Arial", 12, "bold"), foreground="blue")
        self.lbl_score_value.pack(side=tk.LEFT, padx=10)

        self.update_score()

    def get_score_details(self):
        """Puanlamayı hesaplar, her görevin puanını da döndürür."""
        total_score = 0
        # Genel toplam
        for (var, item_data) in self.checkbox_vars:
            if var.get():
                total_score += item_data["points"]
        for (var, item_data) in self.numeric_vars:
            total_score += var.get() * item_data["points_per_item"]

        # Görev bazında puan
        mission_points_dict = {}
        for mission in MISSIONS_DATA:
            m_code = mission["code"]
            m_total = 0
            for item in mission["items"]:
                if item["type"] == "checkbox":
                    for (var, data_ref) in self.checkbox_vars:
                        if data_ref is item and var.get():
                            m_total += data_ref["points"]
                elif item["type"] == "numeric":
                    for (var, data_ref) in self.numeric_vars:
                        if data_ref is item:
                            m_total += var.get() * data_ref["points_per_item"]
            mission_points_dict[m_code] = m_total

        return total_score, mission_points_dict

    def update_score(self, *_):
        """Anlık skor gösterimi."""
        total_score, _ = self.get_score_details()
        self.lbl_score_value.config(text=str(total_score))
        self.big_score_label.config(text=f"Skor: {total_score}")

    def save_score(self):
        """
        Kaydet butonuna basıldığında:
        1) Masaüstünde krustydenemeler.xlsx aç/oluştur,
        2) Her görev için "Mxx Puan" sütunu,
        3) Görev puanı 0 ise ilgili hücreyi KIRMIZI yap.
        """
        total_score, mission_points_dict = self.get_score_details()

        # Masaüstü
        desktop_path = os.path.join(os.path.expanduser("~"), "Desktop")
        filename = "krustydenemeler.xlsx"
        filepath = os.path.join(desktop_path, filename)

        # Excel aç/oluştur
        if os.path.exists(filepath):
            wb = load_workbook(filepath)
            ws = wb.active
        else:
            wb = Workbook()
            ws = wb.active
            ws.title = "Submerged Skor Kaydı"

            # Başlık satırı
            headers = ["Tarih-Saat"]
            for m in MISSIONS_DATA:
                headers.append(f"{m['code']} Puan")
            headers.append("Toplam Puan")
            ws.append(headers)

        # Yeni satır
        now_str = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        row_data = [now_str]

        # Görev puanları
        for m in MISSIONS_DATA:
            m_pts = mission_points_dict[m["code"]]
            row_data.append(m_pts)

        # En sonda Toplam
        row_data.append(total_score)

        # Satırı ekle
        ws.append(row_data)

        # En son eklenen satırın row index'i
        last_row = ws.max_row
        # Sütun sayısı => 1 (Tarih-Saat) + len(MISSIONS_DATA) + 1 (Toplam Puan)

        start_col = 2
        end_col = 1 + len(MISSIONS_DATA)

        # 0 puanlı görev hücrelerini kırmızı yaz
        for col_index in range(start_col, end_col + 1):
            cell = ws.cell(row=last_row, column=col_index)
            if cell.value == 0:
                cell.font = Font(color="FF0000")

        # Dosyayı kaydet
        wb.save(filepath)
        wb.close()

        messagebox.showinfo("Kayıt", "Skor başarıyla kaydedildi!")

    def clear_scores(self):
        """
        Tüm checkbox’ları ve numeric değerleri sıfırlar.
        """
        for (var, item_data) in self.checkbox_vars:
            var.set(False)  # Onay kutusu boş
        for (var, item_data) in self.numeric_vars:
            var.set(0)      # Sayısal giriş 0

        self.update_score()
        messagebox.showinfo("Temizlendi", "Tüm skorlar sıfırlandı!")


def main():
    app = SubmergedScorerApp()
    app.mainloop()

if __name__ == "__main__":
    main()
